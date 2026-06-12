import os
import re
import openpyxl 
from google import genai
from google.genai import types

# GCP 프로젝트 정보 세팅
PROJECT_ID = "gcp-cw-ai-chatbot"
LOCATION = "global" 

client = genai.Client(
    enterprise=True,
    project=PROJECT_ID,
    location=LOCATION
)

def extract_team_name_from_path(file_path: str) -> str:
    """
    실제 구글 드라이브 가상 경로든, 로컬 PC 경로든
    슬래시(/)나 대괄호([ ]) 유무와 상관없이 'OO팀'만 완벽하게 사출하는 무결성 엔진
    """
    # 1. 윈도우 경로인 역슬래시(\)를 슬래시(/)로 통합 치환하여 표준화
    normalized_path = file_path.replace("\\", "/")
    
    # 2. 경로 문자열을 슬래시 기준으로 쪼개서 폴더명 리스트 생성
    path_segments = normalized_path.split("/")
    
    # 3. 가장 하위 폴더(오른쪽)부터 시작해서 역순으로 '팀' 자매품 탐색
    for segment in reversed(path_segments):
        # 혹시 모를 대괄호, 특수문자, 공백을 싹 벗겨내고 순수 텍스트만 추출
        clean_segment = re.sub(r'[\[\]\s]', '', segment)
        
        # 정제된 폴더명이 '팀'으로 끝나면 (예: 총무팀, 인사전략팀, GHR팀 등) 즉시 반환!
        if clean_segment.endswith("팀"):
            return clean_segment
            
    # 4. [최종 방어선] 만약 분할 탐색에 실패했다면 전체 문자열에서 'OO팀' 형태를 정규식으로 통째로 추출
    match = re.search(r'([가-힣\w]+팀)', normalized_path)
    if match:
        return match.group(1).replace("[", "").replace("]", "").strip()
        
    return "공통부서" # 끝까지 매칭이 안 될 경우 빅쿼리 널(Null) 방지용 기본값

def parse_spreadsheet_to_markdown(file_path: str) -> str:
    """Excel/구글시트 변환본 정밀 링크-데이터 파서"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
        md_sheets = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            md_sheets.append(f"## 📊 시트 탭명: {sheet_name}\n")
            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    val = cell.value
                    if val is None:
                        row_values.append("")
                        continue
                    if cell.hyperlink:
                        val = f"[{str(val)}]({cell.hyperlink.target})"
                    elif str(val).startswith("=HYPERLINK"):
                        match = re.search(r'HYPERLINK\(\"([^\"]+)\"\s*,\s*\"([^\"]+)\"\)', str(val), re.IGNORECASE)
                        if match:
                            val = f"[{match.group(2)}]({match.group(1)})"
                    row_values.append(str(val).replace("\n", " "))
                if any(row_values):
                    md_sheets.append("| " + " | ".join(row_values) + " |")
            md_sheets.append("\n")
        return "\n".join(md_sheets)
    except Exception as e:
        print(f"⚠️ [스프레드시트 파서 오작동]: {e}")
        return ""

def parse_document_to_markdown(file_path: str) -> tuple[str, str]:
    """
    [수정완료] 리턴값을 (마크다운결과, 동적추출된팀명) 튜플 형태로 전달하여
    상위 파이프라인에서 빅쿼리 'dept_code' 또는 'dept_name' 컬럼에 즉시 꽂아넣을 수 있도록 개선했습니다.
    """
    # 🌟 프로님이 말씀하신 드라이브 폴더 기반 동적 팀명 추출 엔진 가동!
    extracted_team = extract_team_name_from_path(file_path)
    print(f"🎯 [드라이브 출처 추적 성공] 파일의 소속 부서: '{extracted_team}'")

    # 1. 스프레드시트 검출 및 처리
    if file_path.lower().endswith(('.xlsx', '.xls', '.csv')):
        excel_md = parse_spreadsheet_to_markdown(file_path)
        if excel_md:
            return excel_md, extracted_team

    # 2. Gemini 3.5 Flash 파이프라인 (PDF, 이미지)
    model_name = "gemini-3.5-flash" 
    with open(file_path, "rb") as f:
        document_data = f.read()
    
    mime_type = "application/pdf" if file_path.lower().endswith(".pdf") else "image/png"
    document_part = types.Part.from_bytes(data=document_data, mime_type=mime_type)
    
    prompt = """
    당신은 세계 최고의 문서 구조화(Parsing) 전문가입니다.
    첨부된 문서를 읽고, 내용을 처음부터 끝까지 마크다운(Markdown) 포맷으로 완벽하게 변환하세요.
    문서에 포함된 표(Table)나 셀 내부의 하이퍼링크 URL 주소는 Markdown 표 및 링크 문법으로 100% 복원해야 합니다.
    """
    
    response = client.models.generate_content(
        model=model_name,
        contents=[document_part, prompt]
    )
    
    return response.text, extracted_team

# ==========================================
# 🧪 테스트용 실행 코드 (드라이브 경로 시뮬레이션)
# ==========================================
if __name__ == "__main__":
    # 🧪 일부러 경로에 '총무팀'과 '본사사옥관리'를 가짜로 섞어 넣은 가상 경로 생성
    fake_drive_path = "C:/G_Drive/경영지원본부/총무팀/본사사옥관리/sample_rule.pdf"
    
    # 1. 파일이 실제로 없더라도 팀명 추출이 잘 되는지 먼저 검증!
    detected_team = extract_team_name_from_path(fake_drive_path)
    print("\n" + "="*50)
    print(f"📁 가상 드라이브 입력 경로: {fake_drive_path}")
    print(f"🎯 동적 추출된 출처 부서명: {detected_team} (➔ '총무팀'이 나오면 대성공!)")
    print("="*50 + "\n")